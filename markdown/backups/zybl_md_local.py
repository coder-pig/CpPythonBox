# -*- coding: utf-8 -*-
# !/usr/bin/env python
"""
-------------------------------------------------
   File     : zybl_md_local.py
   Author   : CoderPig
   date     : 2023-10-30 14:21
   Desc     : cmd markdown(作业部落) 数据备份xlsx文件 批量生成本地MD文件的脚本
-------------------------------------------------
"""
import asyncio
import os

from backups_util import md_to_local
from util.file_util import is_dir_existed, search_all_file, write_text_to_file
from openpyxl import load_workbook
from util.logger_util import default_logger
import re

logger = default_logger()
output_dir = os.path.join(os.getcwd(), "output{}{}".format(os.sep, "zybl"))
origin_md_dir = os.path.join(output_dir, "origin_md")  # 生成的原始md文件
local_md_dir = os.path.join(output_dir, "local_md")  # 生成的本地md文件

pic_match_pattern = re.compile(r'(http.*?\.(png|PNG|jpg|JPG|gif|GIF|svg|SVG|webp|awebp))',
                               re.M)  # 匹配图片的正则


# 解析xlsx文件，生成原始md文件
def analysis_xlsx(xlsx_file):
    logger.info("开始生成原始md文件。。。")
    count = 0
    workbook = load_workbook(xlsx_file)
    sheet = workbook['SheetJS']
    # 遍历每一行，min_row → 从第几行开始迭代，values_only=True只获取单元格的值
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if row[0] != "name" and row[1] != "title":
            if row[1] is None:
                logger.info("未检测到标题" + row[2])
            else:
                content = row[2]
                if content is None:
                    logger.info("未检测到文章内容【{}】".format(row[1]))
                    content = ""
                # 需要过滤掉不支持的文件名字符
                count += 1
                md_file = os.path.join(origin_md_dir,
                                       row[1].replace('"', "'")
                                       .replace("?", "？")
                                       .replace("|", "_")
                                       .replace("/", "_")
                                       + ".md")
                logger.info("{}、原始MD文件生成完毕 → {}".format(count, md_file))
                write_text_to_file(content, md_file)
    workbook.close()
    logger.info("共计生成原始MD文件【{}】个 O(∩_∩)O ~".format(count))


def pic_url_func(result):
    """
    从正则匹配结果提取图片URL的方法
    :param result: 正则匹配结果
    :return: 图片URL
    """
    return result[1]


def pic_suffix_func(result):
    """
    从正则匹配结果提取图片后缀的方法
    :param result: 正则匹配结果
    :return: 图片后缀
    """
    return result[2]


def replace_result_func(result, relative_path):
    """
    返回替换后的字符串
    :param result: 正则匹配结果
    :param relative_path: 图片的相对路径
    :return:
    """
    return "{}".format(relative_path)


if __name__ == '__main__':
    is_dir_existed(origin_md_dir)
    is_dir_existed(local_md_dir)
    choose = int(input("请输入要批处理的操作序号：\n1、xlsx生成原始md文件；\n2、原始md文件生成本地md文件：\n"))
    if choose == 1:
        xlsx_files = search_all_file(os.getcwd(), "xlsx")
        if len(xlsx_files) == 0:
            exit("未检测到数据备份的xlsx文件！！！")
        else:
            analysis_xlsx(xlsx_files[0])
    elif choose == 2:
        loop = asyncio.get_event_loop()
        md_to_local(origin_md_dir, local_md_dir, pic_match_pattern, loop, pic_url_func, pic_suffix_func,
                    replace_result_func)
    else:
        exit('错误输入')
