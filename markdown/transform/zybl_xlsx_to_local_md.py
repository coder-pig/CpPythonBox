# -*- coding: utf-8 -*-
# !/usr/bin/env python
"""
-------------------------------------------------
   File     : zybl_xlsx_to_local_md.py
   Author   : CoderPig
   date     : 2023-09-22 17:08
   Desc     : cmd markdown(作业部落) 数据备份xlsx文件 批量生成本地MD文件的脚本
-------------------------------------------------
"""
import asyncio
import os
import re
import time
from functools import partial

import aiofiles
from aiohttp_requests import requests
from openpyxl import load_workbook

origin_md_dir = os.path.join(os.getcwd(), "origin_md")  # 生成的原始md文件
local_md_dir = os.path.join(os.getcwd(), "local_md")  # 生成的本地md文件
pic_url_path_record_list = []  # 存储图片URL和本地图片路径对应关系的列表，用作批量下载图片
pic_match_pattern = re.compile(r'(\]: |\()+(http.*?\.(png|PNG|jpg|JPG|gif|GIF|svg|SVG|webp|awebp))\??(\)?)',
                               re.M)  # 匹配图片的正则
order_set = {i for i in range(1, 500000)}  # 避免图片名重复后缀
default_headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/85.0.4183.121 Safari/537.36 '
}


# 判断目录是否存在，不存在新建
def is_dir_existed(file_path, mkdir=True):
    if mkdir:
        if not os.path.exists(file_path):
            os.makedirs(file_path)
    else:
        return os.path.exists(file_path)


# 把文本写入到文件中
def write_text_to_file(content, file_path, mode="w+"):
    try:
        with open(file_path, mode, encoding='utf-8') as f:
            f.write(content + "\n", )
    except OSError as reason:
        print(str(reason))


# 以文本形式读取文件
def read_file_text_content(file_path):
    if not os.path.exists(file_path):
        return None
    else:
        with open(file_path, 'r+', encoding='utf-8') as f:
            return f.read()


# 扫描特定目录下特定文件后缀，返回文件路径列表
def scan_file_list_by_suffix(file_dir=os.getcwd(), suffix=""):
    return [os.path.join(file_dir, x) for x in os.listdir(file_dir) if x.endswith(suffix)]


# 异步下载图片
async def download_pic(pic_path, url, headers=None):
    try:
        if headers is None:
            headers = default_headers
        if url.startswith("http") | url.startswith("https"):
            if os.path.exists(pic_path):
                print("图片已存在，跳过下载：%s" % pic_path)
            else:
                resp = await requests.get(url, headers=headers)
                print("下载图片：%s" % resp.url)
                if resp is not None:
                    if resp.status != 404:
                        async with aiofiles.open(pic_path, "wb+") as f:
                            await f.write(await resp.read())
                            print("图片下载完毕：%s" % pic_path)
                    else:
                        print("图片不存在：{}".format(url))
        else:
            print("图片链接格式不正确：%s - %s" % (pic_path, url))
    except Exception as e:
        print("下载异常：{}\n{}".format(url, e))


# 解析xlsx文件，生成原始md文件
def analysis_xlsx(xlsx_file):
    workbook = load_workbook(xlsx_file)
    sheet = workbook['SheetJS']
    # 遍历每一行，min_row → 从第几行开始迭代，values_only=True只获取单元格的值
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if row[0] != "name" and row[1] != "title":
            if row[1] is None:
                print("未检测到标题" + row[2])
            else:
                content = row[2]
                if content is None:
                    print("未检测到文章内容【{}】".format(row[1]))
                    content = ""
                # 需要过滤掉不支持的文件名字符
                md_file = os.path.join(origin_md_dir,
                                       row[1].replace('"', "'")
                                       .replace("?", "？")
                                       .replace("|", "_")
                                       .replace("/", "_")
                                       + ".md")
                print("【{}】原始MD文件生成完毕".format(md_file))
                write_text_to_file(content, md_file)
    workbook.close()


# md文件本地化
def md_to_local():
    origin_md_list = scan_file_list_by_suffix(origin_md_dir, "md")
    if len(origin_md_list) == 0:
        exit("未检测到md文件")
    else:
        for origin_md in origin_md_list:
            md_file_name = os.path.basename(origin_md)
            new_md_dir = os.path.join(local_md_dir, md_file_name[:-3])
            new_picture_dir = os.path.join(new_md_dir, "images")
            is_dir_existed(new_md_dir)
            is_dir_existed(new_picture_dir)
            # 生成md文件路径
            new_md_file_path = os.path.join(new_md_dir, md_file_name)
            # 读取md文件内容
            old_content = read_file_text_content(origin_md)
            # 替换原内容
            new_content = pic_match_pattern.sub(partial(pic_to_local, pic_save_dir=new_picture_dir), old_content)
            # 生成新的md文件
            write_text_to_file(new_content, new_md_file_path)
            print("新md文件已生成 → {}".format(new_md_file_path))
        print("所有本地md文件生成完毕！开始批量下载图片文件")
        for pic_url_path_record in pic_url_path_record_list:
            split_list = pic_url_path_record.split("\t")
            loop.run_until_complete(download_pic(split_list[1], split_list[0]))


# 远程图片转换为本地图片
def pic_to_local(match_result, pic_save_dir):
    global pic_url_path_record_list
    print("替换前的图片路径：{}".format(match_result[2]))
    # 生成新的图片名
    img_file_name = "{}_{}.{}".format(int(round(time.time())), order_set.pop(), match_result[3])
    # 拼接图片相对路径(Markdown用到的)
    relative_path = 'images/{}'.format(img_file_name)
    # 拼接图片绝对路径，下载到本地
    absolute_path = os.path.join(pic_save_dir, img_file_name)
    print("替换后的图片路径：{}".format(relative_path))
    pic_url_path_record_list.append("{}\t{}".format(match_result[2], absolute_path))
    # 拼接前后括号()
    return "{}{}{}".format(match_result[1], relative_path, match_result[4])


if __name__ == '__main__':
    is_dir_existed(origin_md_dir)
    is_dir_existed(local_md_dir)
    choose = int(input("请输入要批处理的操作序号：\n1、xlsx生成原始md文件；\n2、原始md文件生成本地md文件：\n"))
    if choose == 1:
        xlsx_files = scan_file_list_by_suffix(suffix="xlsx")
        if len(xlsx_files) == 0:
            exit("未检测到数据备份的xlsx文件！！！")
        else:
            analysis_xlsx(xlsx_files[0])
    elif choose == 2:
        loop = asyncio.get_event_loop()
        md_to_local()
    else:
        exit("错误输入")
